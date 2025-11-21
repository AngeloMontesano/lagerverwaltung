# lagerverwaltung/routes/einstellungen.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from models import db, Kunde, Administratives, Artikel
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
from email.message import EmailMessage
import os
import logging
import smtplib
import json


bp = Blueprint("einstellungen", __name__, url_prefix="/einstellungen")

@bp.route("/save_firmendaten", methods=["POST"])
def save_firmendaten():
    """
    Speichert die Firmendaten aus dem Formular.
    """
    kunde = Kunde.query.first()

    if not kunde:
        kunde = Kunde()

    try:
        kunde.name = request.form.get("name", "")
        kunde.adresse_str = request.form.get("adresse_str", "")
        kunde.adresse_plz = request.form.get("adresse_plz", "")
        kunde.adresse_ort = request.form.get("adresse_ort", "")
        kunde.email = request.form.get("email", "")
        kunde.tel = request.form.get("tel", "")
        kunde.kontakt = request.form.get("kontakt", "")
        kunde.filialnummer = request.form.get("filialnummer", "")
        kunde.steuernummer = request.form.get("steuernummer", "")

        db.session.add(kunde)
        db.session.commit()

        flash("Firmendaten erfolgreich gespeichert!", "success")
        logger.info("Firmendaten wurden gespeichert.")

    except Exception as e:
        logger.error(f"Fehler beim Speichern der Firmendaten: {str(e)}")
        flash("Fehler beim Speichern der Firmendaten!", "danger")

    return redirect(url_for("einstellungen.einstellungen"))

@bp.route("/save_auto_bestellung", methods=["POST"])
def save_auto_bestellung():
    """
    Speichert die Einstellungen für automatische Bestellungen.
    """
    try:
        auto_bestellung = "1" if request.form.get('auto_bestellung') else "0"
        min_bestand = request.form.get('min_bestand', "0")

        Administratives.set_einstellung("auto_bestellung", auto_bestellung)
        Administratives.set_einstellung("auto_bestellung_min", min_bestand)

        db.session.commit()
        flash('Automatische Bestell-Einstellungen gespeichert.', 'success')
        logger.info("Automatische Bestell-Einstellungen gespeichert.")

    except Exception as e:
        logger.error(f"Fehler beim Speichern der automatischen Bestell-Einstellungen: {str(e)}")
        flash("Fehler beim Speichern der Einstellungen!", "danger")

    return redirect(url_for('einstellungen.einstellungen'))

@bp.route("/save_benutzer", methods=["POST"])
def save_benutzer():
    """
    Speichert Benutzerdaten.
    """
    try:
        # Hier könnte die Logik zum Speichern der Benutzer stehen
        flash('Benutzer gespeichert.', 'success')
        logger.info("Benutzer erfolgreich gespeichert.")

    except Exception as e:
        logger.error(f"Fehler beim Speichern der Benutzer: {str(e)}")
        flash("Fehler beim Speichern der Benutzer!", "danger")

    return redirect(url_for('einstellungen.einstellungen'))

@bp.route("/email_empfaenger", methods=["POST"])
def email_empfaenger():
    """
    Speichert die Bestell-Empfänger E-Mail und das Export-Format in der Datenbank.
    """
    try:
        email_empfaenger = request.form.get('email_empfaenger', "").strip()
        export_format = request.form.get('export_format', "csv")

        if email_empfaenger:
            Administratives.set_einstellung("bestell_email_empfaenger", email_empfaenger)
            logger.info(f"Bestell-Empfänger E-Mail gespeichert: {email_empfaenger}")

        Administratives.set_einstellung("export_format", export_format)

        db.session.commit()
        flash('System-Einstellungen gespeichert.', 'success')

    except Exception as e:
        logger.error(f"Fehler beim Speichern der System-Einstellungen: {str(e)}")
        flash("Fehler beim Speichern der System-Einstellungen!", "danger")

    return redirect(url_for('einstellungen.einstellungen'))


    
@bp.route("/", methods=["GET", "POST"])
def einstellungen():
    """
    Zeigt die Firmeneinstellungen und erlaubt Änderungen.
    """
    kunde = Kunde.query.first()
    
    # Bestell-Empfänger E-Mail abrufen
    bestell_email = Administratives.get_einstellung("bestell_email_empfaenger", "")

    # Automatische Bestellungen abrufen
    auto_bestellung = Administratives.get_einstellung("auto_bestellung", "0")
    min_bestand = Administratives.get_einstellung("auto_bestellung_min", "0")

    # Exportformat abrufen
    export_format = Administratives.get_einstellung("export_format", "csv")

    return render_template(
        "einstellungen.html",
        kunde=kunde,
        email_empfaenger=bestell_email,
        auto_bestellung=auto_bestellung,
        min_bestand=min_bestand,
        export_format=export_format
    )

################################
def update_letzte_bestellung():
    """
    Wenn `letzte_bestellung` leer (NULL) ist, wird der Wert von `letzte_aenderung` übernommen.
    """
    try:
        # Update alle Artikel, bei denen letzte_bestellung NULL ist
        Artikel.query.filter(Artikel.letzte_bestellung == None).update(
            {Artikel.letzte_bestellung: Artikel.letzte_aenderung}
        )
        db.session.commit()
        print("✅ Erfolgreich alle Artikel aktualisiert.")
    except Exception as e:
        print(f"❌ Fehler beim Aktualisieren der Artikel: {str(e)}")
        db.session.rollback()

##############################
@bp.route("/send_test_email", methods=["POST"])
def send_test_email():
    """
    Sendet eine Test-E-Mail an die in den Einstellungen hinterlegte Adresse.
    """
    smtp_path = os.path.join("scripts", "smtp.json")
    
    try:
        with open(smtp_path, "r") as file:
            config = json.load(file)
        print("✅ SMTP-Config geladen:", config)
    except FileNotFoundError:
        print(f"❌ Datei nicht gefunden: {smtp_path}")
        flash("SMTP-Konfigurationsdatei nicht gefunden!", "danger")
        return redirect(url_for("einstellungen.einstellungen"))
    except json.JSONDecodeError as e:
        print(f"❌ Fehler beim Parsen der JSON-Datei: {e}")
        flash(f"Fehler beim Parsen der JSON-Datei: {e}", "danger")
        return redirect(url_for("einstellungen.einstellungen"))
    except Exception as e:
        print(f"❌ Unbekannter Fehler beim Laden von smtp.json: {e}")
        flash(f"Unbekannter Fehler beim Laden der SMTP-Konfiguration: {e}", "danger")
        return redirect(url_for("einstellungen.einstellungen"))

    EMAIL_ABSENDER = config.get("EMAIL_ABSENDER")
    EMAIL_PASSWORT = config.get("EMAIL_PASSWORT")
    SMTP_SERVER = config.get("SMTP_SERVER")
    SMTP_PORT = config.get("SMTP_PORT")

    # 📩 **Empfänger aus den Einstellungen laden oder aus Form nehmen**
    email_empfaenger = request.form.get("test_email_empfaenger") or Administratives.get_einstellung("test_email_empfaenger", "angelo@montesano.email")

    msg = EmailMessage()
    msg["Subject"] = "Test-E-Mail von Pfersich Flow"
    msg["From"] = EMAIL_ABSENDER
    msg["To"] = email_empfaenger
    msg.set_content("Dies ist eine Test-E-Mail von Pfersich Flow, um sicherzustellen, dass der SMTP-Server funktioniert.")

    try:
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(EMAIL_ABSENDER, EMAIL_PASSWORT)
            server.send_message(msg)
        
        print(f"✅ Test-E-Mail erfolgreich gesendet an: {email_empfaenger}")
        flash(f"Test-E-Mail wurde erfolgreich gesendet an: {email_empfaenger}", "success")
    except Exception as e:
        print(f"❌ Fehler beim Senden der Test-E-Mail: {str(e)}")
        flash(f"Fehler beim Senden der Test-E-Mail: {str(e)}", "danger")

    return redirect(url_for("einstellungen.einstellungen"))

##############################
# Funktion, die überprüft, ob die Datei eine erlaubte Erweiterung hat
def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'xlsx', 'xls'}  # Erlaubte Dateitypen
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
        


###############################
# Mass Export
@bp.route("/mass_export", methods=["POST"])
def mass_export():
    """
    Exportiert alle Artikel in eine Excel-Datei.
    """
    # Excel-Datei erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Artikel"

    # Spaltenüberschriften definieren
    ws.append(["Artikel-ID", "Name", "Barcode", "Kategorie", "Soll", "Min", "Bestand", 
               "Haltbarkeit", "Charge", "Kunden-ID", "Lagerort", "Preis", "Beschreibung", 
               "Letzte Änderung", "Letzte Bestellung", "Bestellt"])

    # Artikeldaten aus der Datenbank holen
    for artikel in Artikel.query.all():
        ws.append([
            artikel.pf_artikel_id, artikel.name, artikel.ean, artikel.kategorie, 
            artikel.sollbestand, artikel.mindestbestand, artikel.bestand,
            artikel.haltbarkeit, 
            getattr(artikel, 'charge', ''),  # Sicherstellen, dass das 'charge' Feld existiert
            artikel.kunde_id, artikel.lagerort, artikel.preis, 
            getattr(artikel, 'beschreibung', ''),  # Sicherstellen, dass das 'beschreibung' Feld existiert
            artikel.letzte_aenderung.strftime('%Y-%m-%d') if artikel.letzte_aenderung else '',
            getattr(artikel, 'letzte_bestellung', '').strftime('%Y-%m-%d') if artikel.letzte_bestellung else '',
            artikel.bestellt
        ])
    
    # Die Excel-Datei speichern
    filename = "/tmp/artikel_export.xlsx"
    wb.save(filename)

    # Die Excel-Datei zum Herunterladen anbieten
    return send_file(filename, as_attachment=True, download_name="artikel_export.xlsx")


#####################



@bp.route("/import_template", methods=["GET"])
def import_template():
    """
    Erzeugt eine Excel-Datei mit Dummy-Daten als Import-Template.
    """
    # Erstelle ein neues Excel-Arbeitsbuch
    wb = Workbook()
    ws = wb.active
    ws.title = "Artikel Import Template"

    # Dummy-Daten als Überschrift hinzufügen
    headers = [
        "Artikel-ID", "Name", "Barcode", "Kategorie", "Soll", "Min", "Bestand", 
        "Haltbarkeit", "Charge", "Kunden-ID", "Lagerort", "Preis", "Beschreibung", 
        "Letzte Änderung", "Letzte Bestellung", "Bestellt"
    ]
    ws.append(headers)

    # Dummy-Daten hinzufügen
    dummy_data = [
        ["Pfersich Artikel ID 1", "Dummy Artikel", "123456789", "Frische Früchte", 10, 5, 2, 365, "", 1, "Regal A1", 1.2, "Artikel beschreibung", "", "", 0],
        ["Pfersich Artikel ID 2", "Dummy Artikel", "987654321", "Trocken Früchte", 20, 10, 1, 365, "", 1, "Regal A2", 2.0, "Artikel beschreibung", "", "", 0],
        ["Pfersich Artikel ID 3", "Dummy Artikel", "987456123", "Süßungsmittel", 100, 50, 25, 365, "", 1, "Truhe", 3.0, "Artikel beschreibung", "", "", 0]
    ]

    # Dummy-Daten in die Excel-Datei einfügen
    for row in dummy_data:
        ws.append(row)

    # Speichern der Excel-Datei
    filename = "/tmp/import_template.xlsx"
    wb.save(filename)

    # Die Excel-Datei zum Download anbieten
    return send_file(filename, as_attachment=True, download_name="import_template.xlsx")


######################
@bp.route("/mass_import", methods=["POST"])
def mass_import():
    """
    Verarbeitet eine hochgeladene Excel-Datei und nimmt Mass Änderungen vor.
    """
    if "file" not in request.files:
        flash("Keine Datei hochgeladen!", "error")
        return redirect(url_for("einstellungen.einstellungen"))
    
    file = request.files["file"]
    if file.filename == "":
        flash("Keine Datei ausgewählt!", "error")
        return redirect(url_for("einstellungen.einstellungen"))
    
    # Überprüfen, ob die Datei eine erlaubte Erweiterung hat
    if file and allowed_file(file.filename):
        filepath = os.path.join("/tmp", secure_filename(file.filename))
        file.save(filepath)

        # Excel-Datei laden
        wb = load_workbook(filepath)
        ws = wb.active


        # Durch die Excel-Datei iterieren und die Artikel aktualisieren oder hinzufügen
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Entpacken der ersten 13 Werte und der Rest wird in 'extra_values' gespeichert
            pf_artikel_id, name, ean, kategorie, soll, min_bestand, bestand, haltbarkeit, charge, kunde_id, lagerort, preis, beschreibung, *extra_values = row

            # Jetzt kannst du sicherstellen, dass du nur die 13 erwarteten Werte bekommst
            print(f"Artikel {pf_artikel_id} - Name: {name}, EAN: {ean}, Preis: {preis}, Beschreibung: {beschreibung}")

            # Artikel in der Datenbank finden
            artikel = Artikel.query.filter_by(pf_artikel_id=pf_artikel_id).first()
            if artikel:
                # Artikel existiert, also aktualisieren
                artikel.name = name
                artikel.ean = ean
                artikel.kategorie = kategorie
                artikel.sollbestand = soll
                artikel.mindestbestand = min_bestand
                artikel.bestand = bestand
                artikel.haltbarkeit = haltbarkeit
                artikel.charge = charge
                artikel.kunde_id = kunde_id
                artikel.lagerort = lagerort
                artikel.preis = preis
                artikel.beschreibung = beschreibung

                # Wenn letzte_bestellung leer ist, dann setze es auf letzte_aenderung
                if not artikel.letzte_bestellung:
                    artikel.letzte_bestellung = artikel.letzte_aenderung

            else:
                # Artikel existiert nicht, also neu anlegen
                artikel = Artikel(
                    pf_artikel_id=pf_artikel_id,
                    name=name,
                    ean=ean,
                    kategorie=kategorie,
                    sollbestand=soll,
                    mindestbestand=min_bestand,
                    bestand=bestand,
                    haltbarkeit=haltbarkeit,
                    charge=charge,
                    kunde_id=kunde_id,
                    lagerort=lagerort,
                    preis=preis,
                    beschreibung=beschreibung
                )
                db.session.add(artikel)

        db.session.commit()


        flash("Mass Änderungen erfolgreich durchgeführt!", "success")
        return redirect(url_for("einstellungen.einstellungen"))
    
    flash("Ungültige Datei hochgeladen!", "error")
    return redirect(url_for("einstellungen.einstellungen"))